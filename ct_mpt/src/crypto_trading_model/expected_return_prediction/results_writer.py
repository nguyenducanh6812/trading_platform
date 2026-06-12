"""
Results Writer for Expected Return Prediction
===========================================

Writes calculated values back to existing OC analysis files for monitoring.
Adds AR.L columns and prediction columns to the right of existing columns.
"""

import pandas as pd
import numpy as np
import logging
from pathlib import Path
from typing import Dict, Any, Optional, List
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

class WriterError(Exception):
    """Exception raised for results writing errors."""
    pass

class ResultsWriter:
    """
    Writes predicted return results to OC analysis files.
    Preserves existing data structure while adding new calculation columns.
    """
    
    def __init__(self):
        """Initialize the results writer."""
        pass
    
    def write_predictions_to_file(self, enhanced_data: pd.DataFrame, output_path: str, asset_code: str) -> str:
        """
        Write enhanced data with predictions to Excel file.
        
        Args:
            enhanced_data: DataFrame with all original and calculated columns
            output_path: Path to output Excel file
            asset_code: Asset code for naming
            
        Returns:
            Path to the written file
            
        Raises:
            WriterError: If writing fails
        """
        try:
            logger.info(f"Writing prediction results to: {output_path}")
            
            # Ensure output directory exists
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Rename columns for backtester compatibility
            enhanced_data_copy = enhanced_data.copy()
            enhanced_data_copy = self._rename_columns_for_backtester(enhanced_data_copy, asset_code)
            
            # Organize columns for better readability
            organized_data = self._organize_columns(enhanced_data_copy, asset_code)
            
            # Write to Excel file with multiple sheets if needed
            if output_path.suffix.lower() == '.xlsx':
                self._write_to_excel(organized_data, output_path, asset_code)
            else:
                # Convert to Excel if not already
                excel_path = output_path.with_suffix('.xlsx')
                self._write_to_excel(organized_data, excel_path, asset_code)
                output_path = excel_path
            
            # Validate the written file
            self._validate_written_file(output_path, organized_data)
            
            logger.info(f"Results successfully written to: {output_path}")
            return str(output_path)
            
        except Exception as e:
            if isinstance(e, WriterError):
                raise
            else:
                raise WriterError(f"Failed to write prediction results: {str(e)}")
    
    def _rename_columns_for_backtester(self, data: pd.DataFrame, asset_code: str) -> pd.DataFrame:
        """
        Rename columns to match backtester expectations.
        
        Args:
            data: DataFrame with prediction columns
            asset_code: Asset code (e.g., 'BTC', 'ETH')
            
        Returns:
            DataFrame with renamed columns for backtester compatibility
        """
        try:
            # Rename Prd_Return_Arima to Prd_Return_Arima_{asset} for backtester
            if 'Prd_Return_Arima' in data.columns:
                new_column_name = f'Prd_Return_Arima_{asset_code}'
                data = data.rename(columns={'Prd_Return_Arima': new_column_name})
                logger.info(f"Renamed 'Prd_Return_Arima' to '{new_column_name}' for backtester compatibility")
            
            return data
            
        except Exception as e:
            logger.warning(f"Failed to rename columns for backtester: {str(e)}")
            return data  # Return original data if renaming fails
    
    def _organize_columns(self, data: pd.DataFrame, asset_code: str) -> pd.DataFrame:
        """
        Organize columns for better readability.
        Order: Original columns, AR.L columns, Prediction columns
        
        Args:
            data: DataFrame with all columns
            asset_code: Asset code for identifying prediction columns
            
        Returns:
            DataFrame with organized column order
        """
        try:
            # Identify different types of columns
            original_columns = []
            ar_lag_columns = []
            prediction_columns = []
            other_columns = []
            
            # Expected prediction column names
            prediction_column_patterns = ['Prd_Diff_OC', 'Prd_OC', f'Prd_Return_Arima_{asset_code}']
            
            for col in data.columns:
                if col.startswith('Ar.L'):
                    ar_lag_columns.append(col)
                elif col in prediction_column_patterns:
                    prediction_columns.append(col)
                elif col in ['has_complete_ar_lags']:
                    other_columns.append(col)
                else:
                    original_columns.append(col)
            
            # Sort AR lag columns numerically
            ar_lag_columns = sorted(ar_lag_columns, key=lambda x: int(x.split('L')[1]))
            
            # Define the desired column order
            column_order = original_columns + ar_lag_columns + prediction_columns + other_columns
            
            # Reorder the DataFrame
            organized_data = data[column_order]
            
            logger.info(f"Organized columns:")
            logger.info(f"  Original columns: {len(original_columns)}")
            logger.info(f"  AR lag columns: {len(ar_lag_columns)}")
            logger.info(f"  Prediction columns: {len(prediction_columns)}")
            logger.info(f"  Other columns: {len(other_columns)}")
            
            return organized_data
            
        except Exception as e:
            logger.warning(f"Failed to organize columns: {str(e)}")
            return data  # Return original data if organization fails
    
    def _write_to_excel(self, data: pd.DataFrame, output_path: Path, asset_code: str) -> None:
        """
        Write data to Excel file with proper formatting.
        
        Args:
            data: DataFrame to write
            output_path: Path to Excel file
            asset_code: Asset code for sheet naming
            
        Raises:
            WriterError: If writing fails
        """
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Main data sheet
                sheet_name = f"{asset_code}_Analysis_with_Predictions"
                data.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Create summary sheet
                summary_data = self._create_summary_sheet(data, asset_code)
                summary_sheet_name = f"{asset_code}_Prediction_Summary"
                summary_data.to_excel(writer, sheet_name=summary_sheet_name, index=False)
                
                logger.info(f"Created Excel sheets: {sheet_name}, {summary_sheet_name}")
            
            # Apply additional formatting if needed
            self._apply_excel_formatting(output_path, data)
            
        except Exception as e:
            raise WriterError(f"Failed to write Excel file: {str(e)}")
    
    def _create_summary_sheet(self, data: pd.DataFrame, asset_code: str) -> pd.DataFrame:
        """
        Create a summary sheet with key statistics.
        
        Args:
            data: Full data DataFrame
            asset_code: Asset code
            
        Returns:
            DataFrame with summary statistics
        """
        try:
            summary_rows = []
            
            # Basic information
            summary_rows.append(['Asset Code', asset_code])
            summary_rows.append(['Total Rows', len(data)])
            summary_rows.append(['Analysis Date', pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')])
            summary_rows.append(['', ''])  # Blank row
            
            # AR Lag Information
            ar_columns = [col for col in data.columns if col.startswith('Ar.L')]
            summary_rows.append(['AR Lag Order', len(ar_columns)])
            summary_rows.append(['AR Lag Columns', ', '.join(ar_columns[:5]) + ('...' if len(ar_columns) > 5 else '')])
            summary_rows.append(['', ''])  # Blank row
            
            # Prediction Statistics
            prediction_columns = ['Prd_Diff_OC', 'Prd_OC', f'Prd_Return_Arima_{asset_code}']
            
            for col in prediction_columns:
                if col in data.columns:
                    col_data = data[col].dropna()
                    if len(col_data) > 0:
                        summary_rows.append([f'{col} - Valid Count', len(col_data)])
                        summary_rows.append([f'{col} - Mean', f'{col_data.mean():.6f}'])
                        summary_rows.append([f'{col} - Std Dev', f'{col_data.std():.6f}'])
                        summary_rows.append([f'{col} - Min', f'{col_data.min():.6f}'])
                        summary_rows.append([f'{col} - Max', f'{col_data.max():.6f}'])
                        summary_rows.append(['', ''])  # Blank row
            
            # Data Quality
            summary_rows.append(['Data Quality Metrics', ''])
            for col in ['Demean_Diff_OC', 'OC', 'Open_Price'] + prediction_columns:
                if col in data.columns:
                    null_count = data[col].isnull().sum()
                    null_pct = (null_count / len(data)) * 100
                    summary_rows.append([f'{col} - Null Count', f'{null_count} ({null_pct:.1f}%)'])
            
            # Convert to DataFrame
            summary_df = pd.DataFrame(summary_rows, columns=['Metric', 'Value'])
            
            return summary_df
            
        except Exception as e:
            logger.warning(f"Failed to create summary sheet: {str(e)}")
            # Return basic summary if detailed creation fails
            return pd.DataFrame([
                ['Asset Code', asset_code],
                ['Total Rows', len(data)],
                ['Analysis Date', pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')]
            ], columns=['Metric', 'Value'])
    
    def _apply_excel_formatting(self, file_path: Path, data: pd.DataFrame) -> None:
        """
        Apply basic formatting to Excel file.
        
        Args:
            file_path: Path to Excel file
            data: DataFrame data for context
        """
        try:
            # Load the workbook
            workbook = load_workbook(file_path)
            
            # Format main data sheet
            for sheet_name in workbook.sheetnames:
                if 'Analysis_with_Predictions' in sheet_name:
                    sheet = workbook[sheet_name]
                    
                    # Freeze first row (headers)
                    sheet.freeze_panes = sheet['A2']
                    
                    # Auto-adjust column widths (basic)
                    for column in sheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 20)  # Cap at 20 characters
                        sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save the formatted workbook
            workbook.save(file_path)
            workbook.close()
            
            logger.info("Applied basic Excel formatting")
            
        except Exception as e:
            logger.warning(f"Failed to apply Excel formatting: {str(e)}")
            # Don't raise error - formatting is optional
    
    def _validate_written_file(self, file_path: Path, expected_data: pd.DataFrame) -> None:
        """
        Validate the written file contains expected data.
        
        Args:
            file_path: Path to written file
            expected_data: Expected DataFrame
            
        Raises:
            WriterError: If validation fails
        """
        try:
            if not file_path.exists():
                raise WriterError(f"Output file was not created: {file_path}")
            
            # Read back the file to validate
            written_data = pd.read_excel(file_path, sheet_name=0)
            
            # Basic validation
            if len(written_data) != len(expected_data):
                raise WriterError(f"Row count mismatch: expected {len(expected_data)}, got {len(written_data)}")
            
            # Check for key columns - look for asset-specific prediction column
            base_columns = ['Prd_Diff_OC', 'Prd_OC']
            missing_columns = [col for col in base_columns if col not in written_data.columns]
            
            # Check for asset-specific prediction column
            asset_prediction_columns = [col for col in written_data.columns if col.startswith('Prd_Return_Arima_')]
            if not asset_prediction_columns:
                missing_columns.append('Prd_Return_Arima_{asset}')
            
            if missing_columns:
                logger.error(f"Available columns in written file: {list(written_data.columns)}")
                raise WriterError(f"Missing prediction columns in written file: {missing_columns}")
            
            logger.info(f"File validation passed. Found prediction columns: {base_columns + asset_prediction_columns}")
            
        except Exception as e:
            if isinstance(e, WriterError):
                raise
            else:
                raise WriterError(f"File validation failed: {str(e)}")
    
    def backup_original_file(self, file_path: str) -> Optional[str]:
        """
        Create a backup of the original file before overwriting.
        
        Args:
            file_path: Path to file to backup
            
        Returns:
            Path to backup file, or None if backup failed
        """
        try:
            original_path = Path(file_path)
            if not original_path.exists():
                return None
            
            # Create backup filename with timestamp
            timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
            backup_name = f"{original_path.stem}_backup_{timestamp}{original_path.suffix}"
            backup_path = original_path.parent / backup_name
            
            # Copy the file
            import shutil
            shutil.copy2(original_path, backup_path)
            
            logger.info(f"Created backup: {backup_path}")
            return str(backup_path)
            
        except Exception as e:
            logger.warning(f"Failed to create backup: {str(e)}")
            return None