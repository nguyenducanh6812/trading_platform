"""
ARIMA Forecast Reporter
======================

File: crypto_trading_model/reporting/arima_forecast_reporter.py

This module provides specialized reporting for ARIMA forecasting results with
easy monitoring and tracking capabilities.

Creates two sheets:
1. Time Series Data: Date, Return, Demean, AR.1-AR.3, MA.1-MA.5, E, Prd demean, Prd Return
2. Master Data: Mean value, ARIMA pdq, AR/MA coefficients, model statistics

Usage:
    reporter = ArimaForecastReporter(logger)
    reporter.export_forecast_monitoring_data(
        forecasting_results, 
        output_path="arima_monitoring.xlsx"
    )
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime
import csv

from ..custom_logging.logger import CustomLogger
from ..arima.arima_forecaster import ArimaForecaster, ForecastResult, MultiForecastResult
from ..arima.arima_optimizer import ArimaResult

class ArimaForecastReporter:
    """
    Specialized reporter for ARIMA forecasting results with monitoring focus.
    
    This class creates structured reports that match your Excel format for
    easy tracking and monitoring of ARIMA forecasting performance.
    """
    
    def __init__(self, logger: CustomLogger):
        """
        Initialize the ARIMA Forecast Reporter.
        
        Args:
            logger: Custom logger instance
        """
        self.logger = logger
        self.logger.info("ArimaForecastReporter initialized")

    def export_forecast_monitoring_data(
        self,
        arima_result: ArimaResult,
        forecaster: ArimaForecaster,
        historical_data: pd.Series,
        forecast_results: Optional[MultiForecastResult] = None,
        output_path: str = "arima_monitoring.xlsx",
        include_forecasts: bool = True,
        forecast_steps: int = 5
    ) -> Dict[str, pd.DataFrame]:
        """
        Export complete ARIMA monitoring data with time series and master data.
        
        Args:
            arima_result: ARIMA optimization result
            forecaster: Configured ARIMA forecaster
            historical_data: Historical return series
            forecast_results: Optional pre-generated forecast results
            output_path: Output file path (.xlsx or .csv)
            include_forecasts: Whether to include future forecasts
            forecast_steps: Number of forecast steps if generating new forecasts
            
        Returns:
            Dictionary with 'time_series' and 'master_data' DataFrames
        """
        try:
            self.logger.info(f"Generating ARIMA monitoring data for export to {output_path}")
            
            # Generate time series data sheet
            time_series_df = self._create_time_series_sheet(
                arima_result, forecaster, historical_data, 
                forecast_results, include_forecasts, forecast_steps
            )
            
            # Generate master data sheet
            master_data_df = self._create_master_data_sheet(
                arima_result, forecaster, historical_data, time_series_df
            )
            
            # Export based on file extension
            if output_path.endswith('.xlsx'):
                self._export_to_excel(time_series_df, master_data_df, output_path)
            elif output_path.endswith('.csv'):
                self._export_to_csv_files(time_series_df, master_data_df, output_path)
            else:
                # Default to Excel
                excel_path = str(Path(output_path).with_suffix('.xlsx'))
                self._export_to_excel(time_series_df, master_data_df, excel_path)
                self.logger.warning(f"Unknown extension, exported to Excel: {excel_path}")
            
            self.logger.info(f"ARIMA monitoring data exported successfully")
            
            return {
                'time_series': time_series_df,
                'master_data': master_data_df
            }
            
        except Exception as e:
            error_msg = f"Failed to export ARIMA monitoring data: {e}"
            self.logger.error(error_msg)
            raise Exception(error_msg)

    def _create_time_series_sheet(
        self,
        arima_result: ArimaResult,
        forecaster: ArimaForecaster,
        historical_data: pd.Series,
        forecast_results: Optional[MultiForecastResult],
        include_forecasts: bool,
        forecast_steps: int
    ) -> pd.DataFrame:
        """
        Create the time series sheet with columns:
        Date, Return, Demean, AR.1, AR.2, AR.3, MA.1, MA.2, MA.3, MA.4, MA.5, E, Prd demean, Prd Return
        """
        self.logger.info("Creating time series data sheet...")
        
        # Initialize the DataFrame structure
        time_series_data = []
        
        # Get model parameters
        p, d, q = arima_result.best_order
        mean_value = forecaster.mean_value
        
        # Process historical data
        historical_returns = historical_data.values
        historical_dates = historical_data.index
        
        # Calculate demeaned values
        demeaned_values = historical_returns - mean_value
        
        # Initialize AR and MA histories
        ar_values_history = [0.0] * max(p, 3)  # At least 3 AR columns
        ma_values_history = [0.0] * max(q, 5)  # At least 5 MA columns  
        error_history = [0.0] * max(q, 5)
        
        # Process each historical period
        for i in range(len(historical_returns)):
            date = historical_dates[i]
            return_val = historical_returns[i]
            demean_val = demeaned_values[i]
            
            # Get AR values (lagged demeaned values)
            ar_1 = ar_values_history[-1] if len(ar_values_history) > 0 else 0.0
            ar_2 = ar_values_history[-2] if len(ar_values_history) > 1 else 0.0
            ar_3 = ar_values_history[-3] if len(ar_values_history) > 2 else 0.0
            
            # Get MA values (lagged error terms)
            ma_1 = error_history[-1] if len(error_history) > 0 else 0.0
            ma_2 = error_history[-2] if len(error_history) > 1 else 0.0
            ma_3 = error_history[-3] if len(error_history) > 2 else 0.0
            ma_4 = error_history[-4] if len(error_history) > 3 else 0.0
            ma_5 = error_history[-5] if len(error_history) > 4 else 0.0
            
            # Calculate predicted demean using ARIMA coefficients
            prd_demean = self._calculate_predicted_demean(
                [ar_1, ar_2, ar_3], [ma_1, ma_2, ma_3, ma_4, ma_5], 
                arima_result.significant_coefficients
            )
            
            # Calculate predicted return
            prd_return = prd_demean + mean_value
            
            # Calculate error term
            error_term = demean_val - prd_demean
            
            # Create row data
            row_data = {
                'Date': date,
                'Return': return_val,
                'Demean': demean_val,
                'AR.1': ar_1,
                'AR.2': ar_2,
                'AR.3': ar_3,
                'MA.1': ma_1,
                'MA.2': ma_2,
                'MA.3': ma_3,
                'MA.4': ma_4,
                'MA.5': ma_5,
                'E': error_term,
                'Prd demean': prd_demean,
                'Prd Return': prd_return
            }
            
            time_series_data.append(row_data)
            
            # Update histories for next iteration
            ar_values_history.append(demean_val)
            error_history.append(error_term)
            
            # Keep histories at reasonable length
            if len(ar_values_history) > 10:
                ar_values_history.pop(0)
            if len(error_history) > 10:
                error_history.pop(0)
        
        # Add future forecasts if requested
        if include_forecasts:
            if forecast_results is None:
                # Generate new forecasts
                try:
                    forecast_results = forecaster.predict_multiple_steps(forecast_steps)
                except Exception as e:
                    self.logger.warning(f"Could not generate forecasts: {e}")
                    forecast_results = None
            
            if forecast_results and forecast_results.forecasts:
                time_series_data.extend(
                    self._add_forecast_rows(
                        forecast_results, historical_dates[-1], mean_value
                    )
                )
        
        # Convert to DataFrame
        df = pd.DataFrame(time_series_data)
        
        # Ensure proper column order
        column_order = [
            'Date', 'Return', 'Demean', 'AR.1', 'AR.2', 'AR.3', 
            'MA.1', 'MA.2', 'MA.3', 'MA.4', 'MA.5', 'E', 'Prd demean', 'Prd Return'
        ]
        
        df = df.reindex(columns=column_order, fill_value=0.0)
        
        self.logger.info(f"Created time series sheet with {len(df)} rows")
        return df

    def _calculate_predicted_demean(
        self, 
        ar_values: List[float], 
        ma_values: List[float],
        significant_coefficients: Dict[str, float]
    ) -> float:
        """Calculate predicted demean using SUMPRODUCT of AR/MA values and coefficients."""
        
        prd_demean = 0.0
        
        # Add AR contributions
        for i, ar_val in enumerate(ar_values, 1):
            coeff_name = f'ar.L{i}'
            if coeff_name in significant_coefficients:
                contribution = ar_val * significant_coefficients[coeff_name]
                prd_demean += contribution
        
        # Add MA contributions  
        for i, ma_val in enumerate(ma_values, 1):
            coeff_name = f'ma.L{i}'
            if coeff_name in significant_coefficients:
                contribution = ma_val * significant_coefficients[coeff_name]
                prd_demean += contribution
        
        return prd_demean

    def _add_forecast_rows(
        self,
        forecast_results: MultiForecastResult,
        last_historical_date: pd.Timestamp,
        mean_value: float
    ) -> List[Dict]:
        """Add forecast rows to time series data."""
        
        forecast_rows = []
        
        for i, forecast in enumerate(forecast_results.forecasts):
            # Generate future date
            forecast_date = last_historical_date + pd.Timedelta(days=i+1)
            
            # Create forecast row (Return will be actual when available)
            row_data = {
                'Date': forecast_date,
                'Return': np.nan,  # To be filled with actual data
                'Demean': np.nan,  # To be calculated from actual return
                'AR.1': forecast.ar_values[0] if len(forecast.ar_values) > 0 else 0.0,
                'AR.2': forecast.ar_values[1] if len(forecast.ar_values) > 1 else 0.0,
                'AR.3': forecast.ar_values[2] if len(forecast.ar_values) > 2 else 0.0,
                'MA.1': forecast.ma_values[0] if len(forecast.ma_values) > 0 else 0.0,
                'MA.2': forecast.ma_values[1] if len(forecast.ma_values) > 1 else 0.0,
                'MA.3': forecast.ma_values[2] if len(forecast.ma_values) > 2 else 0.0,
                'MA.4': forecast.ma_values[3] if len(forecast.ma_values) > 3 else 0.0,
                'MA.5': forecast.ma_values[4] if len(forecast.ma_values) > 4 else 0.0,
                'E': forecast.error_term,
                'Prd demean': forecast.predict_demean,
                'Prd Return': forecast.predict_return
            }
            
            forecast_rows.append(row_data)
        
        return forecast_rows

    def _create_master_data_sheet(
        self,
        arima_result: ArimaResult,
        forecaster: ArimaForecaster,
        historical_data: pd.Series,
        time_series_df: pd.DataFrame
    ) -> pd.DataFrame:
        """
        Create master data sheet with:
        Mean value, ARIMA pdq, AR coefficients, MA coefficients, model statistics
        """
        self.logger.info("Creating master data sheet...")
        
        master_data = []
        
        # Basic model information
        p, d, q = arima_result.best_order
        master_data.extend([
            {'Parameter', 'Value', 'Description'},
            {'=== MODEL INFORMATION ===', '', ''},
            {'ARIMA Order (p,d,q)', f'({p},{d},{q})', 'Autoregressive, Differencing, Moving Average orders'},
            {'AIC', arima_result.best_aic, 'Akaike Information Criterion'},
            {'Forecast Ready', arima_result.forecast_ready, 'Model ready for forecasting'},
            {'Significant Coefficients', len(arima_result.significant_coefficients), 'Number of significant parameters'},
            {'', '', ''}
        ])
        
        # Data information
        master_data.extend([
            {'=== DATA INFORMATION ===', '', ''},
            {'Historical Periods', len(historical_data), 'Number of historical observations'},
            {'Mean Value', forecaster.mean_value, 'Mean of historical returns'},
            {'Data Start Date', historical_data.index[0], 'First observation date'},
            {'Data End Date', historical_data.index[-1], 'Last observation date'},
            {'Return Std Dev', historical_data.std(), 'Standard deviation of returns'},
            {'Return Min', historical_data.min(), 'Minimum return'},
            {'Return Max', historical_data.max(), 'Maximum return'},
            {'', '', ''}
        ])
        
        # ARIMA coefficients
        master_data.append({'=== ARIMA COEFFICIENTS ===', '', ''})
        
        # AR coefficients
        ar_coeffs_found = False
        for coeff_name, coeff_value in arima_result.significant_coefficients.items():
            if coeff_name.startswith('ar.L'):
                lag = coeff_name.split('ar.L')[1]
                master_data.append({f'AR.{lag}', coeff_value, f'Autoregressive coefficient lag {lag}'})
                ar_coeffs_found = True
        
        if not ar_coeffs_found:
            master_data.append({'AR Coefficients', 'None significant', 'No significant AR terms found'})
        
        master_data.append({'', '', ''})
        
        # MA coefficients
        ma_coeffs_found = False
        for coeff_name, coeff_value in arima_result.significant_coefficients.items():
            if coeff_name.startswith('ma.L'):
                lag = coeff_name.split('ma.L')[1]
                master_data.append({f'MA.{lag}', coeff_value, f'Moving average coefficient lag {lag}'})
                ma_coeffs_found = True
        
        if not ma_coeffs_found:
            master_data.append({'MA Coefficients', 'None significant', 'No significant MA terms found'})
        
        master_data.append({'', '', ''})
        
        # Other significant coefficients
        other_coeffs = {k: v for k, v in arima_result.significant_coefficients.items() 
                       if not k.startswith('ar.L') and not k.startswith('ma.L')}
        if other_coeffs:
            master_data.append({'=== OTHER COEFFICIENTS ===', '', ''})
            for coeff_name, coeff_value in other_coeffs.items():
                master_data.append({coeff_name, coeff_value, f'Other model parameter: {coeff_name}'})
            master_data.append({'', '', ''})
        
        # Model diagnostics
        if hasattr(arima_result, 'model_diagnostics') and arima_result.model_diagnostics:
            diagnostics = arima_result.model_diagnostics
            master_data.extend([
                {'=== MODEL DIAGNOSTICS ===', '', ''},
                {'Log Likelihood', diagnostics.get('log_likelihood', 'N/A'), 'Model log likelihood'},
                {'BIC', diagnostics.get('bic', 'N/A'), 'Bayesian Information Criterion'},
                {'In-Sample RMSE', diagnostics.get('in_sample_rmse', 'N/A'), 'Root Mean Square Error'},
                {'In-Sample MAE', diagnostics.get('in_sample_mae', 'N/A'), 'Mean Absolute Error'},
                {'Ljung-Box p-value', diagnostics.get('ljung_box_pvalue', 'N/A'), 'Residual autocorrelation test'},
                {'Residuals White Noise', diagnostics.get('residuals_are_white_noise', 'N/A'), 'Residuals pass white noise test'},
                {'', '', ''}
            ])
        
        # Performance metrics from time series if available
        if len(time_series_df) > 0:
            # Calculate forecast accuracy metrics
            actual_returns = time_series_df['Return'].dropna()
            predicted_returns = time_series_df['Prd Return'].dropna()
            
            if len(actual_returns) > 0 and len(predicted_returns) > 0:
                # Align data
                min_len = min(len(actual_returns), len(predicted_returns))
                if min_len > 1:
                    actual_subset = actual_returns.iloc[:min_len]
                    predicted_subset = predicted_returns.iloc[:min_len]
                    
                    mae = np.mean(np.abs(actual_subset - predicted_subset))
                    mse = np.mean((actual_subset - predicted_subset) ** 2)
                    rmse = np.sqrt(mse)
                    
                    master_data.extend([
                        {'=== FORECAST PERFORMANCE ===', '', ''},
                        {'Mean Absolute Error', mae, 'Average absolute forecast error'},
                        {'Root Mean Square Error', rmse, 'Root mean square forecast error'},
                        {'Forecast Accuracy %', max(0, (1 - mae/actual_subset.std()) * 100), 'Relative forecast accuracy'},
                        {'', '', ''}
                    ])
        
        # Generation metadata
        master_data.extend([
            {'=== REPORT METADATA ===', '', ''},
            {'Generated Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'Report generation timestamp'},
            {'Model Order String', f'ARIMA{arima_result.best_order}', 'Full model specification'},
            {'Total Parameters', len(arima_result.significant_coefficients), 'Number of model parameters'},
            {'Report Version', '1.0', 'Reporter version'}
        ])
        
        # Convert to DataFrame - handle the tuple unpacking properly
        df_data = []
        for item in master_data:
            if isinstance(item, (list, tuple)) and len(item) >= 3:
                df_data.append({
                    'Parameter': item[0],
                    'Value': item[1], 
                    'Description': item[2]
                })
            elif isinstance(item, dict):
                df_data.append(item)
            else:
                # Handle malformed entries
                df_data.append({
                    'Parameter': str(item) if not isinstance(item, (list, tuple)) else str(item[0]),
                    'Value': '',
                    'Description': ''
                })
        
        df = pd.DataFrame(df_data)
        
        self.logger.info(f"Created master data sheet with {len(df)} parameters")
        return df

    def _export_to_excel(self, time_series_df: pd.DataFrame, master_data_df: pd.DataFrame, 
                        output_path: str):
        """Export both sheets to Excel file."""
        
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Export time series data
                time_series_df.to_excel(writer, sheet_name='Time Series Data', index=False)
                
                # Export master data
                master_data_df.to_excel(writer, sheet_name='Master Data', index=False)
                
                # Format the sheets for better readability
                self._format_excel_sheets(writer, time_series_df, master_data_df)
            
            self.logger.info(f"Exported ARIMA monitoring data to Excel: {output_path}")
            
        except ImportError:
            self.logger.warning("openpyxl not available, falling back to CSV export")
            csv_path = str(Path(output_path).with_suffix('.csv'))
            self._export_to_csv_files(time_series_df, master_data_df, csv_path)
        except Exception as e:
            self.logger.error(f"Excel export failed: {e}")
            raise

    def _export_to_csv_files(self, time_series_df: pd.DataFrame, master_data_df: pd.DataFrame,
                           base_path: str):
        """Export to separate CSV files (since CSV doesn't support multiple sheets)."""
        
        base_path = Path(base_path)
        base_name = base_path.stem
        base_dir = base_path.parent
        
        # Export time series data
        time_series_path = base_dir / f"{base_name}_time_series.csv"
        time_series_df.to_csv(time_series_path, index=False, float_format='%.8f')
        
        # Export master data
        master_data_path = base_dir / f"{base_name}_master_data.csv"
        master_data_df.to_csv(master_data_path, index=False)
        
        self.logger.info(f"Exported ARIMA monitoring data to CSV files:")
        self.logger.info(f"  Time Series: {time_series_path}")
        self.logger.info(f"  Master Data: {master_data_path}")

    def _format_excel_sheets(self, writer, time_series_df: pd.DataFrame, 
                           master_data_df: pd.DataFrame):
        """Apply formatting to Excel sheets for better readability."""
        
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Get workbook and worksheets
            workbook = writer.book
            ts_sheet = writer.sheets['Time Series Data']
            master_sheet = writer.sheets['Master Data']
            
            # Format time series sheet
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in ts_sheet[1]:  # First row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in ts_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ts_sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Format master data sheet similarly
            for cell in master_sheet[1]:  # First row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Highlight section headers in master data
            section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            for row in master_sheet.iter_rows(min_row=2):
                if row[0].value and str(row[0].value).startswith('==='):
                    for cell in row:
                        cell.fill = section_fill
                        cell.font = Font(bold=True)
            
            self.logger.info("Applied Excel formatting")
            
        except ImportError:
            self.logger.warning("openpyxl styling not available, exported without formatting")
        except Exception as e:
            self.logger.warning(f"Excel formatting failed: {e}")

    def create_monitoring_summary(self, time_series_df: pd.DataFrame) -> Dict[str, Any]:
        """Create a summary for monitoring dashboard."""
        
        # Calculate key metrics
        total_periods = len(time_series_df)
        forecast_periods = time_series_df['Return'].isna().sum()
        historical_periods = total_periods - forecast_periods
        
        # Forecast accuracy (where both actual and predicted are available)
        valid_data = time_series_df.dropna(subset=['Return', 'Prd Return'])
        
        summary = {
            'total_periods': total_periods,
            'historical_periods': historical_periods, 
            'forecast_periods': forecast_periods,
            'data_start_date': time_series_df['Date'].iloc[0],
            'data_end_date': time_series_df['Date'].iloc[-1],
            'last_historical_date': time_series_df[time_series_df['Return'].notna()]['Date'].iloc[-1] if historical_periods > 0 else None
        }
        
        if len(valid_data) > 0:
            actual_returns = valid_data['Return']
            predicted_returns = valid_data['Prd Return']
            
            errors = np.abs(actual_returns - predicted_returns)
            summary.update({
                'forecast_mae': errors.mean(),
                'forecast_rmse': np.sqrt(np.mean(errors ** 2)),
                'forecast_accuracy_pct': max(0, (1 - errors.mean() / actual_returns.std()) * 100),
                'max_forecast_error': errors.max(),
                'min_forecast_error': errors.min()
            })
        
        return summary

# Example usage and testing functions
def test_arima_forecast_reporter():
    """Test the ARIMA Forecast Reporter with sample data."""
    
    logger = CustomLogger('ArimaReporterTest')
    reporter = ArimaForecastReporter(logger)
    
    # This would normally use real ARIMA results
    # For testing, we'll create mock data
    logger.info("ARIMA Forecast Reporter test would require real ARIMA results")
    logger.info("Integrate with your existing ARIMA analysis to test this component")
    
    return reporter

if __name__ == "__main__":
    test_arima_forecast_reporter()
